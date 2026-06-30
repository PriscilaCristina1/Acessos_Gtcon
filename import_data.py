import sqlite3
import openpyxl
import re

DB_PATH = "acessos.db"
EXCEL_PATH = "Acessos GTCON.xlsx"

SHEET_CONFIG = {
    "ACESSOS SERVIDOR": {
        "columns": ["usuario_exact", "usuario_gtcon", "senha_padrao", "ultimo_criado", "observacao"]
    },
    "VAGOS": {
        "columns": ["departamento", "email", "onedrive", "anydesk", "senha_anydesk", "maquina", "observacao"]
    },
    "ENCAMINHADOS": {
        "columns": ["de_email", "para_email", "observacao"]
    },
    "NIT": {
        "columns": ["responsavel", "anydesk", "email", "senha_padrao_anydesk", "onedrive", "senha_maquina", "observacao"]
    },
    "DIRETORIA": {
        "columns": ["colaborador", "anydesk", "email", "senha_padrao_anydesk", "onedrive", "senha_maquina", "observacao"]
    },
    "DP": {
        "columns": ["colaborador", "anydesk", "email", "senha_padrao_anydesk", "onedrive", "certificado_gtcon", "maquina", "departamento", "observacao"]
    },
    "FISCAL": {
        "columns": ["colaborador", "anydesk", "email", "senha_padrao_anydesk", "onedrive", "certificado_gtcon", "maquina", "observacao"]
    },
    "CONTABIL": {
        "columns": ["colaborador", "anydesk", "email", "senha_padrao_anydesk", "onedrive", "certificado_gtcon", "maquina", "observacao"]
    },
    "IMPLANTAÇÃO": {
        "columns": ["colaborador", "anydesk", "email", "senha_padrao_anydesk", "onedrive", "certificado_gtcon", "maquina", "observacao"]
    },
    "COMPLIANCE": {
        "columns": ["colaborador", "anydesk", "email", "senha_padrao_anydesk", "onedrive", "certificado_gtcon", "maquina", "observacao"]
    },
    "LEGALIZAÇÃO": {
        "columns": ["colaborador", "anydesk", "email", "senha_padrao_anydesk", "onedrive", "certificado_gtcon", "maquina", "observacao"]
    },
    "TRIBUTÁRIO": {
        "columns": ["colaborador", "anydesk", "email", "senha_padrao_anydesk", "onedrive", "certificado_gtcon", "maquina", "observacao"]
    },
    "CS": {
        "columns": ["colaborador", "anydesk", "email", "senha_padrao_anydesk", "onedrive", "certificado_gtcon", "maquina", "observacao"]
    },
    "MARKETING": {
        "columns": ["colaborador", "anydesk", "email", "senha_padrao_anydesk", "onedrive", "certificado_gtcon", "maquina", "observacao"]
    },
    "COMERCIAL": {
        "columns": ["colaborador", "anydesk", "email", "senha_padrao_anydesk", "onedrive", "certificado_gtcon", "maquina", "observacao"]
    },
    "DIVERSOS": {
        "columns": ["descricao", "anydesk", "email", "senha_padrao_anydesk", "onedrive", "senha_maquina", "observacao"]
    },
    "TROCA DE E-MAILS": {
        "columns": ["responsavel", "email", "novo_email", "departamento", "observacao"]
    },
    "INF NOTEBOOK": {
        "columns": ["dispositivo", "modelo", "serie", "nome_dispositivo", "processador", "id_dispositivo", "id_produto", "colaborador", "anydesk", "observacao"]
    },
    "CERTIFICADO NAYRA": {
        "columns": ["instalado_por", "dia", "observacao"]
    }
}


def normalize_sheet_name(name):
    return name.strip().upper().replace(" ", "_").replace("Ç", "C").replace("Ã", "A").replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U").replace("Â", "A").replace("Ê", "E").replace("Õ", "O")


def clean_value(val):
    if val is None:
        return ""
    val = str(val).strip()
    val = val.replace("\u2013", "-").replace("\u2014", "-")
    val = val.replace("\u2018", "'").replace("\u2019", "'")
    val = val.replace("\u201c", '"').replace("\u201d", '"')
    val = val.replace("\u00a0", " ").replace("\ufffd", "")
    return val


def is_header_or_empty(row_values):
    text = " ".join(str(v) for v in row_values if v is not None).strip()
    if not text:
        return True
    headers = {"colaborador", "anydesk", "e-mail", "email", "senha", "responsavel", "dispositivo",
               "modelo", "nome", "processador", "departamento", "instalado", "vago", "de", "para",
               "usuario", "maquina", "certificado", "onedrive", "dispositivo"}
    first = str(row_values[0]).strip().lower() if row_values[0] else ""
    if first in headers or first.endswith(":"):
        return True
    return False


def extract_sheet_data(ws, sheet_name):
    rows_data = []
    in_subsection = False
    subsection_name = ""
    current_sub_rows = []
    sub_sections = {}
    all_main_rows = []

    for row in ws.iter_rows(min_row=1, values_only=True):
        values = [clean_value(v) for v in row]

        if not any(v for v in values):
            if subsection_name and current_sub_rows:
                sub_sections[subsection_name] = current_sub_rows
                current_sub_rows = []
                subsection_name = ""
            continue

        first_val = values[0].strip().upper() if values[0] else ""

        if first_val and not first_val.startswith("HTTP") and not first_val.startswith("WWW"):
            if first_val.isupper() and len(first_val) > 2 and not re.match(r'^[A-Z0-9._%+-]+@', first_val):
                if any(v.isupper() for v in values if v and len(v.strip()) > 3):
                    maybe_header = first_val in {"COLABORADOR", "RESPONSÁVEL", "DISPOSITIVO", "EMAIL",
                                                 "E-MAIL", "ANYDESK", "FISCAL", "VAGOS", "DP",
                                                 "COMPLIANCE", "CONTABIL", "TRIBUTARIO", "DIRETORIA",
                                                 "USUÁRIO EXACT", "USUÁRIO GTCON", "SENHA PADRÃO",
                                                 "INSTALADO POR MIM", "DIA", "DE", "PARA"}
                    if maybe_header or (values[0] and "COLABORADOR" in values[0] or "ANYDESK" in values[0]):
                        pass
                    elif values[0] != values[0].strip():
                        pass
                    elif len(first_val) > 3:
                        subsection_name = values[0]
                        in_subsection = True
                        continue

        if in_subsection:
            current_sub_rows.append(values)
        else:
            all_main_rows.append(values)

    if subsection_name and current_sub_rows:
        sub_sections[subsection_name] = current_sub_rows

    if sub_sections:
        result = []
        for sub_name, sub_rows in sub_sections.items():
            for sr in sub_rows:
                if not is_header_or_empty(sr):
                    result.append((sub_name, sr))
        return result

    result = []
    for r in all_main_rows:
        if not is_header_or_empty(r):
            result.append(("", r))
    return result


def create_database():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS sheets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_name TEXT NOT NULL,
            sheet_key TEXT NOT NULL UNIQUE,
            display_order INTEGER DEFAULT 0
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_id INTEGER NOT NULL,
            departamento TEXT DEFAULT '',
            colaborador TEXT DEFAULT '',
            anydesk TEXT DEFAULT '',
            email TEXT DEFAULT '',
            senha_padrao_anydesk TEXT DEFAULT '',
            onedrive TEXT DEFAULT '',
            certificado_gtcon TEXT DEFAULT '',
            maquina TEXT DEFAULT '',
            senha_maquina TEXT DEFAULT '',
            usuario_exact TEXT DEFAULT '',
            usuario_gtcon TEXT DEFAULT '',
            senha_padrao TEXT DEFAULT '',
            responsavel TEXT DEFAULT '',
            dispositivo TEXT DEFAULT '',
            modelo TEXT DEFAULT '',
            serie TEXT DEFAULT '',
            nome_dispositivo TEXT DEFAULT '',
            processador TEXT DEFAULT '',
            id_dispositivo TEXT DEFAULT '',
            id_produto TEXT DEFAULT '',
            instalado_por TEXT DEFAULT '',
            dia TEXT DEFAULT '',
            descricao TEXT DEFAULT '',
            de_email TEXT DEFAULT '',
            para_email TEXT DEFAULT '',
            novo_email TEXT DEFAULT '',
            observacao TEXT DEFAULT '',
            FOREIGN KEY (sheet_id) REFERENCES sheets (id)
        )
    """)

    conn.commit()
    return conn


def clean_sheet_name_for_db(name):
    name = name.strip()
    name_mapping = {
        "ACESSOS SERVIDOR": "ACESSOS_SERVIDOR",
        "TROCA DE E-MAILS": "TROCA_DE_EMAILS",
        "ENCAMINHADOS": "ENCAMINHADOS",
        "CERTIFICADO NAYRA": "CERTIFICADO_NAYRA",
        "INF NOTEBOOK": "INF_NOTEBOOK",
    }
    for k, v in name_mapping.items():
        if name.strip().upper() == k.upper():
            return v

    normalized = name.strip().upper()
    normalized = normalized.replace(" ", "_").replace("Ç", "C").replace("Ã", "A")
    normalized = normalized.replace("Á", "A").replace("É", "E").replace("Í", "I")
    normalized = normalized.replace("Ó", "O").replace("Ú", "U").replace("Â", "A")
    normalized = normalized.replace("Ê", "E").replace("Õ", "O")
    return normalized


def smart_extract(ws, sheet_name):
    config = SHEET_CONFIG.get(sheet_name.strip(), {})
    columns = config.get("columns", ["colaborador", "anydesk", "email", "observacao"])
    data = []

    rows = list(ws.iter_rows(values_only=True))
    header_row_idx = None

    for idx, row in enumerate(rows):
        vals = [clean_value(v) for v in row]
        txt = " ".join(v.lower() for v in vals if v)
        if "colaborador" in txt and ("anydesk" in txt or "email" in txt or "e-mail" in txt):
            header_row_idx = idx
            break

    if sheet_name.strip().upper() == "ACESSOS SERVIDOR":
        return extract_acessos_servidor(rows)
    if sheet_name.strip().upper() == "VAGOS":
        return extract_vagos(rows)
    if sheet_name.strip().upper() == "ENCAMINHADOS":
        return extract_encaminhados(rows)
    if sheet_name.strip().upper() == "TROCA DE E-MAILS":
        return extract_troca_emails(rows)
    if sheet_name.strip().upper() == "CERTIFICADO NAYRA":
        return extract_certificado_nayra(rows)
    if sheet_name.strip().upper() == "INF NOTEBOOK":
        return extract_inf_notebook(rows)
    if sheet_name.strip().upper() == "NIT":
        return extract_nit(rows)
    if sheet_name.strip().upper() == "DIRETORIA":
        return extract_diretoria(rows)
    if sheet_name.strip().upper() == "DIVERSOS":
        return extract_diversos(rows)

    return extract_generic_department(rows, header_row_idx, columns, sheet_name)


def extract_acessos_servidor(rows):
    data = []
    for row in rows:
        vals = [clean_value(v) for v in row]
        if not any(v for v in vals):
            continue
        first = vals[0].upper() if vals[0] else ""
        if first in {"USUÁRIO EXACT", "USUÁRIO GTCON", "SENHA PADRÃO", "", "ULTIMO CRIADO", "USUÁRIO"}:
            if not vals[1] or vals[1].upper() in {"ADM", "USUÁRIO GTCON", ""}:
                continue
        if vals[0] and vals[0].startswith("GTCON"):
            data.append({
                "usuario_exact": vals[0] if vals[0].startswith("GTCON") else "",
                "usuario_gtcon": vals[1] if vals[1] and not vals[1].startswith("GTCON") else "",
                "senha_padrao": vals[2] if vals[2] and not vals[2].startswith(("GTCON", "ÚLTIMO")) else "",
            })
        elif len(vals) > 6 and vals[6] and vals[6].startswith("GTCON"):
            data.append({
                "usuario_exact": vals[6] if vals[6].startswith("GTCON") else "",
                "usuario_gtcon": vals[7] if vals[7] and not vals[7].startswith("GTCON") else "",
                "senha_padrao": vals[8] if len(vals) > 8 and vals[8] and not vals[8].startswith(("GTCON", "SENHA")) else "",
            })
    return data


def extract_vagos(rows):
    data = []
    current_dept = ""
    for row in rows:
        vals = [clean_value(v) for v in row]
        if not any(v for v in vals):
            continue
        first = vals[0].strip().upper() if vals[0] else ""
        if first in {"FISCAL", "COMPLIANCE", "DP", "CONTABIL", "TRIBUTARIO", "TRIBUTÁRIO"}:
            current_dept = vals[0].strip()
            continue
        if first in {"VAGOS", "ONEDRIVE", ""}:
            continue
        if vals[0] and "@" in vals[0]:
            data.append({
                "departamento": current_dept,
                "email": vals[0],
                "onedrive": vals[1] if len(vals) > 1 else "",
                "anydesk": vals[2] if len(vals) > 2 else "",
                "senha_anydesk": vals[3] if len(vals) > 3 else "",
                "maquina": vals[4] if len(vals) > 4 else "",
            })
    return data


def extract_encaminhados(rows):
    data = []
    for row in rows:
        vals = [clean_value(v) for v in row]
        if not any(v for v in vals):
            continue
        first = vals[0].upper().strip() if vals[0] else ""
        if first in {"DE", "PARA", ""}:
            continue
        if vals[0] and "@" in vals[0]:
            data.append({
                "de_email": vals[0],
                "para_email": vals[1] if len(vals) > 1 and vals[1] else "",
            })
    return data


def extract_troca_emails(rows):
    data = []
    current_dept = ""
    header_found = False
    for row in rows:
        vals = [clean_value(v) for v in row]
        if not any(v for v in vals):
            continue
        first = vals[0].strip().upper() if vals[0] else ""
        if first in {"RESPONSÁVEL", "RESPONSAVEL"} and "E-MAIL" in " ".join(v.upper() for v in vals if v):
            header_found = True
            continue
        if first in {"ESTAGIÁRIOS", "ESTAGIARIOS", "IMPLANTAÇÃO", "IMPLANTACAO", "CS"}:
            current_dept = vals[0].strip()
            header_found = False
            continue
        if header_found and vals[0] and vals[0] != "-":
            data.append({
                "responsavel": vals[0],
                "email": vals[1] if len(vals) > 1 else "",
                "novo_email": vals[2] if len(vals) > 2 else "",
                "departamento": current_dept,
            })
        elif not header_found and vals[0] and "@" in vals[0]:
            data.append({
                "responsavel": "",
                "email": vals[0],
                "novo_email": vals[1] if len(vals) > 1 else "",
                "departamento": current_dept,
            })
    return data


def extract_certificado_nayra(rows):
    data = []
    for row in rows:
        vals = [clean_value(v) for v in row]
        if not any(v for v in vals):
            continue
        first = vals[0].upper().strip() if vals[0] else ""
        if first in {"INSTALADO POR MIM", "DIA", ""}:
            continue
        if vals[0] and vals[0] != "-":
            data.append({"instalado_por": vals[0], "dia": vals[1] if len(vals) > 1 else ""})
    return data


def extract_inf_notebook(rows):
    data = []
    for row in rows:
        vals = [clean_value(v) for v in row]
        if not any(v for v in vals):
            continue
        first = vals[0].upper().strip() if vals[0] else ""
        if first in {"DISPOSITIVO", ""}:
            continue
        if vals[0] and vals[0] != "-" and not first.startswith("DISPOSITIVO"):
            data.append({
                "dispositivo": vals[0],
                "modelo": vals[1] if len(vals) > 1 else "",
                "serie": vals[2] if len(vals) > 2 else "",
                "nome_dispositivo": vals[3] if len(vals) > 3 else "",
                "processador": vals[4] if len(vals) > 4 else "",
                "id_dispositivo": vals[5] if len(vals) > 5 else "",
                "id_produto": vals[6] if len(vals) > 6 else "",
                "colaborador": vals[7] if len(vals) > 7 else "",
                "anydesk": vals[8] if len(vals) > 8 else "",
            })
    return data


def extract_nit(rows):
    data = []
    for row in rows:
        vals = [clean_value(v) for v in row]
        if not any(v for v in vals):
            continue
        first = vals[0].upper().strip() if vals[0] else ""
        if first in {"RESPONSÁVEL", "RESPONSAVEL", ""}:
            continue
        if vals[0] and vals[0] != "-":
            data.append({
                "responsavel": vals[0],
                "anydesk": vals[1] if len(vals) > 1 else "",
                "email": vals[2] if len(vals) > 2 else "",
                "senha_padrao_anydesk": vals[3] if len(vals) > 3 else "",
                "onedrive": vals[4] if len(vals) > 4 else "",
                "senha_maquina": vals[5] if len(vals) > 5 else "",
            })
    return data


def extract_diretoria(rows):
    data = []
    for row in rows:
        vals = [clean_value(v) for v in row]
        if not any(v for v in vals):
            continue
        first = vals[0].upper().strip() if vals[0] else ""
        if first in {"DIRETORIA", ""}:
            continue
        if first == "COLABORADOR" or first == "ANYDESK" or first == "E-MAIL":
            continue
        if vals[0] and vals[0] != "-":
            data.append({
                "colaborador": vals[0],
                "anydesk": vals[1] if len(vals) > 1 else "",
                "email": vals[2] if len(vals) > 2 else "",
                "senha_padrao_anydesk": vals[3] if len(vals) > 3 else "",
                "onedrive": vals[4] if len(vals) > 4 else "",
                "senha_maquina": vals[5] if len(vals) > 5 else "",
            })
    return data


def extract_diversos(rows):
    data = []
    for row in rows:
        vals = [clean_value(v) for v in row]
        if not any(v for v in vals):
            continue
        first = vals[0].upper().strip() if vals[0] else ""
        if first in {"EMAIL", "ANYDESK", "E-MAIL", ""}:
            continue
        if vals[0] and vals[0] != "-":
            data.append({
                "descricao": vals[0],
                "anydesk": vals[1] if len(vals) > 1 else "",
                "email": vals[2] if len(vals) > 2 else "",
                "senha_padrao_anydesk": vals[3] if len(vals) > 3 else "",
                "onedrive": vals[4] if len(vals) > 4 else "",
                "senha_maquina": vals[5] if len(vals) > 5 else "",
            })
    return data


def extract_generic_department(rows, header_row_idx, columns, sheet_name):
    if header_row_idx is None:
        return []

    data_rows = rows[header_row_idx + 1:]
    data = []
    current_sub_dept = ""

    for row in data_rows:
        vals = [clean_value(v) for v in row]
        if not any(v for v in vals):
            current_sub_dept = ""
            continue

        first = vals[0].strip().upper() if vals[0] else ""
        if first == sheet_name.strip().upper():
            continue
        if first == "COLABORADOR" and ("ANYDESK" in " ".join(v.upper() for v in vals if v)):
            continue

        colab_col = 0
        anydesk_col = 1
        email_col = 2
        senha_col = 3
        onedrive_col = 4
        cert_col = 5
        maquina_col = 6

        record = {
            "colaborador": vals[colab_col] if len(vals) > colab_col else "",
            "anydesk": vals[anydesk_col] if len(vals) > anydesk_col else "",
            "email": vals[email_col] if len(vals) > email_col else "",
            "senha_padrao_anydesk": vals[senha_col] if len(vals) > senha_col else "",
            "onedrive": vals[onedrive_col] if len(vals) > onedrive_col else "",
            "certificado_gtcon": vals[cert_col] if len(vals) > cert_col else "",
            "maquina": vals[maquina_col] if len(vals) > maquina_col else "",
        }

        if record["colaborador"] and record["colaborador"] != "-":
            data.append(record)

    return data


def import_all():
    conn = create_database()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM records")
    cursor.execute("DELETE FROM sheets")
    conn.commit()

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    order = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        key = clean_sheet_name_for_db(sheet_name)
        cursor.execute(
            "INSERT OR IGNORE INTO sheets (sheet_name, sheet_key, display_order) VALUES (?, ?, ?)",
            (sheet_name.strip(), key, order)
        )
        cursor.execute("SELECT id FROM sheets WHERE sheet_key = ?", (key,))
        sheet_id = cursor.fetchone()[0]

        records = smart_extract(ws, sheet_name)

        inserted = 0
        for rec in records:
            cursor.execute("""
                INSERT INTO records (
                    sheet_id, departamento, colaborador, anydesk, email,
                    senha_padrao_anydesk, onedrive, certificado_gtcon, maquina,
                    senha_maquina, usuario_exact, usuario_gtcon, senha_padrao,
                    responsavel, dispositivo, modelo, serie, nome_dispositivo,
                    processador, id_dispositivo, id_produto, instalado_por, dia,
                    descricao, de_email, para_email, novo_email, observacao
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                sheet_id,
                rec.get("departamento", ""),
                rec.get("colaborador", ""),
                rec.get("anydesk", ""),
                rec.get("email", ""),
                rec.get("senha_padrao_anydesk", ""),
                rec.get("onedrive", ""),
                rec.get("certificado_gtcon", ""),
                rec.get("maquina", ""),
                rec.get("senha_maquina", ""),
                rec.get("usuario_exact", ""),
                rec.get("usuario_gtcon", ""),
                rec.get("senha_padrao", ""),
                rec.get("responsavel", ""),
                rec.get("dispositivo", ""),
                rec.get("modelo", ""),
                rec.get("serie", ""),
                rec.get("nome_dispositivo", ""),
                rec.get("processador", ""),
                rec.get("id_dispositivo", ""),
                rec.get("id_produto", ""),
                rec.get("instalado_por", ""),
                rec.get("dia", ""),
                rec.get("descricao", ""),
                rec.get("de_email", ""),
                rec.get("para_email", ""),
                rec.get("novo_email", ""),
                rec.get("observacao", ""),
            ))
            inserted += 1

        print(f"  {sheet_name}: {inserted} registros importados")
        order += 1

    conn.commit()
    conn.close()
    wb.close()
    print(f"\nTotal de {order} abas importadas com sucesso!")


if __name__ == "__main__":
    import_all()
